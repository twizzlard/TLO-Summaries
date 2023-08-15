import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import zipfile
import os
import tempfile
import streamlit as st

def process_excel(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    # Check headings
    headings = ["Index", "PROBLEM", "RESEARCH", "SOLUTION", "KEYS TO SUCCESS"]
    for idx, heading in enumerate(headings):
        assert sheet.cell(row=1, column=idx+1).value == heading, f"Expected {heading}, but found {sheet.cell(row=1, column=idx+1).value}"

    with tempfile.TemporaryDirectory() as pdf_dir:
        for row_idx in range(2, sheet.max_row + 1):
            index = sheet.cell(row=row_idx, column=1).value
            contents = [sheet.cell(row=row_idx, column=col_idx+1).value for col_idx in range(1, len(headings))]

            pdf_file = f"{pdf_dir}/Entry_{index}.pdf"
            doc = SimpleDocTemplate(pdf_file, pagesize=letter)
            styles = getSampleStyleSheet()
            styles["Normal"].fontSize = 10
            styles["Normal"].leading = 12  # Line height
            content_list = []
            content_list.append(Paragraph(f"Entry# {index}", styles["Normal"]))

            for idx, text_content in enumerate(contents):
                content_list.append(Paragraph(f"{headings[idx+1]}:", styles["Normal"]))
                content_list.append(Paragraph(str(text_content), styles["Normal"]))

            content_list.append(Paragraph("SCORING", styles["Normal"]))
            content_list.append(Paragraph("Research ______/30pts max", styles["Normal"]))
            content_list.append(Paragraph("Solution ______/40pts max", styles["Normal"]))
            content_list.append(Paragraph("Uniqueness ______/20pts max", styles["Normal"]))
            content_list.append(Paragraph("Professionalism ______/10pts max", styles["Normal"]))
            content_list.append(Paragraph("Total Score: ___________", styles["Normal"]))
            content_list.append(Paragraph("NOTES:", styles["Normal"]))

            doc.build(content_list)

        zip_file = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
        with zipfile.ZipFile(zip_file, 'w') as zipf:
            for root, _, files in os.walk(pdf_dir):
                for file in files:
                    zipf.write(os.path.join(root, file), file)

        return zip_file.name

# Streamlit app
st.title("Excel to PDF Converter")
uploaded_file = st.file_uploader("Drag and drop your Excel file here:", type=["xlsx"])

if uploaded_file is not None:
    with st.spinner("Processing..."):
        zip_path = process_excel(uploaded_file)
    st.success("Processing complete!")

    with open(zip_path, 'rb') as file:
        zip_bytes = file.read()
        st.download_button(
            label="Download ZIP file",
            data=zip_bytes,
            file_name="entries.zip",
            mime="application/zip",
        )
